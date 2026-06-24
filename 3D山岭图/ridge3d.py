#!/usr/bin/env python3
"""
============================================================
 3D 山岭图生成器 — 只需提供数据 CSV，即刻生成交互式 HTML
============================================================
用法:
  python ridge3d.py data.csv
  python ridge3d.py data.csv --title "我的研究" --author "张三" --colorscheme ocean
  python ridge3d.py data.xlsx --sheet "Sheet1"

CSV格式 (UTF-8):
  name,value,ci_low,ci_high,weight,group
  研究A,0.5,0.2,0.8,25.0,亚组1
  研究B,0.3,-0.1,0.7,30.0,亚组2
  ...

作者: 马佳宇
版本: 1.0
"""
import argparse, sys, os, json
import numpy as np
import plotly.graph_objects as go


# ============================================================
# 内置配色方案
# ============================================================
COLORSCHEMES = {
    "forest": {
        "surface": [
            [0.0, 'rgb(3, 10, 3)'],
            [0.1, 'rgb(5, 35, 10)'],
            [0.25, 'rgb(0, 80, 30)'],
            [0.45, 'rgb(10, 140, 70)'],
            [0.65, 'rgb(40, 190, 140)'],
            [0.82, 'rgb(70, 215, 220)'],
            [1.0, 'rgb(110, 230, 255)'],
        ],
        "name": "Dark Green -> Sky Blue (Forest)",
        "bg": "white",
    },
    "ocean": {
        "surface": [
            [0.0, 'rgb(2, 5, 30)'],
            [0.15, 'rgb(5, 25, 70)'],
            [0.35, 'rgb(10, 60, 130)'],
            [0.55, 'rgb(30, 120, 200)'],
            [0.75, 'rgb(60, 180, 240)'],
            [1.0, 'rgb(130, 230, 255)'],
        ],
        "name": "Deep Sea -> Bright Blue (Ocean)",
        "bg": "white",
    },
    "sunset": {
        "surface": [
            [0.0, 'rgb(40, 5, 20)'],
            [0.15, 'rgb(80, 15, 30)'],
            [0.35, 'rgb(160, 50, 40)'],
            [0.55, 'rgb(220, 120, 60)'],
            [0.75, 'rgb(250, 180, 100)'],
            [1.0, 'rgb(255, 230, 180)'],
        ],
        "name": "Dark Red -> Gold (Sunset)",
        "bg": "white",
    },
    "glacier": {
        "surface": [
            [0.0, 'rgb(20, 20, 60)'],
            [0.2, 'rgb(40, 60, 120)'],
            [0.4, 'rgb(80, 130, 200)'],
            [0.6, 'rgb(160, 200, 240)'],
            [0.8, 'rgb(210, 235, 255)'],
            [1.0, 'rgb(240, 250, 255)'],
        ],
        "name": "Deep Purple -> Ice White (Glacier)",
        "bg": "white",
    },
    "lava": {
        "surface": [
            [0.0, 'rgb(5, 5, 5)'],
            [0.15, 'rgb(40, 10, 5)'],
            [0.35, 'rgb(140, 30, 10)'],
            [0.55, 'rgb(230, 100, 20)'],
            [0.75, 'rgb(255, 180, 40)'],
            [1.0, 'rgb(255, 240, 120)'],
        ],
        "name": "Black -> Lava Gold (Volcano)",
        "bg": "white",
    },
    "aurora": {
        "surface": [
            [0.0, 'rgb(5, 15, 30)'],
            [0.2, 'rgb(20, 60, 80)'],
            [0.4, 'rgb(40, 140, 120)'],
            [0.6, 'rgb(100, 210, 140)'],
            [0.8, 'rgb(180, 240, 180)'],
            [1.0, 'rgb(230, 255, 230)'],
        ],
        "name": "Night -> Aurora Green (Aurora)",
        "bg": "white",
    },
    "golden": {
        "surface": [
            [0.0, 'rgb(15, 10, 3)'],
            [0.12, 'rgb(50, 35, 5)'],
            [0.25, 'rgb(120, 90, 15)'],
            [0.4, 'rgb(200, 160, 30)'],
            [0.55, 'rgb(240, 210, 60)'],
            [0.7, 'rgb(180, 200, 100)'],
            [0.85, 'rgb(80, 180, 220)'],
            [1.0, 'rgb(50, 140, 255)'],
        ],
        "name": "Gold -> Blue (Yellow-Blue)",
        "bg": "white",
    },
}

# 默认分组颜色
DEFAULT_GROUP_COLORS = [
    "#2196F3", "#FF9800", "#F44336", "#9C27B0",
    "#4CAF50", "#00BCD4", "#E91E63", "#795548",
    "#607D8B", "#FF5722",
]


# ============================================================
# 数据Loaded
# ============================================================
def load_csv(path):
    """从 CSV Loaded数据，自动检测分隔符"""
    import csv
    rows = []
    with open(path, 'r', encoding='utf-8-sig') as f:
        # 尝试检测分隔符
        sample = f.read(4096)
        f.seek(0)
        sniffer = csv.Sniffer()
        dialect = sniffer.sniff(sample)
        reader = csv.DictReader(f, dialect=dialect)
        for row in reader:
            rows.append(row)
    return rows


def load_excel(path, sheet=None):
    """从 Excel Loaded数据"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for h, v in zip(headers, row):
            d[str(h)] = v if v is not None else ''
        if any(v for v in d.values()):
            rows.append(d)
    return rows


def parse_data(rows, args):
    """Parsed数据行，自动匹配列名"""
    # 自动检测列名 (大小写不敏感，中英文兼容)
    col_map = {
        'name': ['name', 'study', 'title', '名称', '研究', '标题', 'label'],
        'value': ['value', 'smd', 'md', 'effect', 'es', 'or', 'rr', 'hr', '值', '效应量', '影响因子'],
        'ci_low': ['ci_low', 'ci_lower', 'lower', 'lo', 'cilo', 'ci下限', '下限'],
        'ci_high': ['ci_high', 'ci_upper', 'upper', 'hi', 'cihi', 'ci上限', '上限'],
        'weight': ['weight', 'w', '权重', 'weight(%)', 'weight%', 'wt'],
        'group': ['group', 'subgroup', 'category', '分组', '亚组', '类别', 'grp'],
    }
    if not rows:
        sys.exit("Error: 数据为空")

    headers = [h.lower().strip() for h in rows[0].keys()]
    mapping = {}
    for key, candidates in col_map.items():
        for c in candidates:
            if c in headers:
                mapping[key] = list(rows[0].keys())[headers.index(c)]
                break
        if key not in mapping:
            if key == 'ci_low' and 'value' in mapping:
                mapping['ci_low'] = None  # CI可选
            elif key == 'ci_high' and 'value' in mapping:
                mapping['ci_high'] = None
            elif key == 'weight':
                mapping['weight'] = None  # weight可选
            elif key == 'group':
                mapping['group'] = None   # group可选
            elif key == 'name':
                sys.exit("Error: 找不到名称列。需要: name/study/title/名称 等")
            elif key == 'value':
                sys.exit("Error: 找不到数值列。需要: value/smd/md/effect/值 等")

    data = []
    for row in rows:
        name = str(row.get(mapping['name'], '')).strip()
        if not name:
            continue
        try:
            val = float(row.get(mapping['value'], 0))
        except (ValueError, TypeError):
            continue

        ci_low = None
        ci_high = None
        weight = 1.0
        group = ''

        if mapping.get('ci_low') and mapping['ci_low'] in row:
            try:
                ci_low = float(row[mapping['ci_low']])
            except (ValueError, TypeError):
                pass
        if mapping.get('ci_high') and mapping['ci_high'] in row:
            try:
                ci_high = float(row[mapping['ci_high']])
            except (ValueError, TypeError):
                pass
        if mapping.get('weight') and mapping['weight'] in row:
            try:
                weight = float(row[mapping['weight']])
            except (ValueError, TypeError):
                pass
        if mapping.get('group') and mapping['group'] in row:
            group = str(row[mapping['group']]).strip()

        if ci_low is None:
            ci_low = val - abs(val) * 0.3
        if ci_high is None:
            ci_high = val + abs(val) * 0.3
        if weight is None or weight <= 0:
            weight = 0.1

        data.append((name, val, ci_low, ci_high, weight, group))

    return data


# ============================================================
# 图表生成
# ============================================================
def build_ridge(data, args):
    """构建 3D 山岭图"""
    n = len(data)
    if n == 0:
        sys.exit("Error: 无有效数据行")

    # 确定配色
    cs = COLORSCHEMES.get(args.colorscheme, COLORSCHEMES["forest"])

    # 确定分组颜色
    groups = list(dict.fromkeys([d[5] for d in data if d[5]]))
    if not groups:
        groups = ["Default"]
    group_colors = {}
    for i, g in enumerate(groups):
        group_colors[g] = DEFAULT_GROUP_COLORS[i % len(DEFAULT_GROUP_COLORS)]

    # Y轴范围
    all_y = [d[1] for d in data]
    all_ci_low = [d[2] for d in data]
    all_ci_high = [d[3] for d in data]
    y_pad = (max(all_ci_high) - min(all_ci_low)) * 0.15
    y_min = min(all_ci_low) - y_pad
    y_max = max(all_ci_high) + y_pad

    # 网格
    y_grid = np.linspace(y_min, y_max, 400)
    x_grid = np.linspace(-0.5, n - 0.5, n * 20)
    X, Y = np.meshgrid(x_grid, y_grid)
    Z = np.zeros_like(X)

    # 叠加高斯山脊
    for i, (name, val, ci_low, ci_high, weight, group) in enumerate(data):
        center = val
        ci_width = max(ci_high - ci_low, 0.01)
        sigma = ci_width / 4.0
        sigma = max(sigma, 0.08)
        # 非线性高度
        w = max(weight, 0.1)
        amplitude = (w / 100.0) ** 0.55 * 2.2
        x_sigma = 0.25

        for j, xi in enumerate(x_grid):
            x_weight = np.exp(-0.5 * ((xi - i) / x_sigma) ** 2)
            if x_weight < 0.01:
                continue
            ridge = amplitude * np.exp(-0.5 * ((y_grid - center) / sigma) ** 2)
            Z[:, j] += ridge * x_weight

    # ========== 构建图表 ==========
    fig = go.Figure()

    fig.add_trace(go.Surface(
        x=X, y=Y, z=Z,
        colorscale=cs["surface"],
        opacity=0.85,
        showscale=True,
        colorbar=dict(title="密度", x=1.02, thickness=20, len=0.6),
        contours=dict(
            z=dict(show=True, usecolormap=True,
                   highlightcolor="white", project=dict(z=False), width=1),
        ),
        lighting=dict(ambient=0.4, diffuse=0.8, fresnel=0.3,
                       specular=0.5, roughness=0.3),
        lightposition=dict(x=3000, y=5000, z=8000),
        name="山岭曲面",
    ))

    # 数据点 + CI线
    for i, (name, val, ci_low, ci_high, weight, group) in enumerate(data):
        color = group_colors.get(group, "#757575")
        w = max(weight, 0.1)
        z_val = (w / 100.0) ** 0.55 * 2.2

        fig.add_trace(go.Scatter3d(
            x=[i], y=[val], z=[z_val + 0.03],
            mode='markers+text',
            marker=dict(size=w * 0.12 + 3, color=color,
                         line=dict(color='white', width=1.5)),
            text=[name],
            textposition='top center',
            textfont=dict(size=9, color='black'),
            hovertext=[f"<b>{name}</b><br>Value: {val:.2f}<br>95%CI: [{ci_low:.2f}, {ci_high:.2f}]<br>Weight: {weight:.1f}%<br>Group: {group}"],
            hoverinfo='text',
            showlegend=False,
        ))

        ci_y = np.linspace(ci_low, ci_high, 40)
        fig.add_trace(go.Scatter3d(
            x=np.full_like(ci_y, i), y=ci_y, z=np.full_like(ci_y, 0.01),
            mode='lines', line=dict(color=color, width=3),
            hoverinfo='skip', showlegend=False,
        ))
        for ep in [ci_low, ci_high]:
            fig.add_trace(go.Scatter3d(
                x=[i], y=[ep], z=[0.02],
                mode='markers',
                marker=dict(size=3, color=color, symbol='diamond'),
                hoverinfo='skip', showlegend=False,
            ))

    # 分组标签
    group_ranges = {}
    for i, (_, _, _, _, _, group) in enumerate(data):
        g = group if group else "Default"
        if g not in group_ranges:
            group_ranges[g] = [i, i]
        group_ranges[g][1] = i

    for g, (s, e) in group_ranges.items():
        mid = (s + e) / 2
        color = group_colors.get(g, "#757575")
        fig.add_trace(go.Scatter3d(
            x=[mid], y=[y_min - (y_max - y_min) * 0.05], z=[0],
            mode='text',
            text=[g],
            textfont=dict(size=12, color=color, family='Arial Black'),
            hoverinfo='skip', showlegend=False,
        ))

    # 图例
    for g, color in group_colors.items():
        fig.add_trace(go.Scatter3d(
            x=[None], y=[None], z=[None],
            mode='markers', marker=dict(size=10, color=color),
            name=g,
        ))

    # 标题
    title_text = f"<b>{args.title}</b>"
    if args.subtitle:
        title_text += f"<br><sub>{args.subtitle}</sub>"
    if args.author:
        title_text += f"<br><sub>—— {args.author}</sub>"

    fig.update_layout(
        title=dict(text=title_text, x=0.5, xanchor='center',
                    font=dict(size=18, family='Arial')),
        scene=dict(
            xaxis=dict(
                title=args.xlabel,
                tickmode='array',
                tickvals=list(range(n)),
                ticktext=[d[0] for d in data],
                tickfont=dict(size=8),
                range=[-1, n],
                gridcolor='rgba(200,200,200,0.3)',
                backgroundcolor='rgba(240,240,245,0.5)',
            ),
            yaxis=dict(
                title=args.ylabel,
                range=[y_min, y_max],
                gridcolor='rgba(200,200,200,0.3)',
                backgroundcolor='rgba(240,240,245,0.5)',
            ),
            zaxis=dict(
                title='权重 (高度)',
                range=[0, 2.5],
                gridcolor='rgba(200,200,200,0.3)',
                backgroundcolor='rgba(240,240,245,0.5)',
            ),
            camera=dict(eye=dict(x=1.6, y=-1.8, z=1.2), up=dict(x=0, y=0, z=1)),
            aspectmode='manual',
            aspectratio=dict(x=1.2, y=1.0, z=0.7),
            bgcolor=cs["bg"],
        ),
        margin=dict(l=0, r=0, t=100, b=0),
        hovermode='closest',
    )

    return fig


# ============================================================
# 主入口
# ============================================================
def main():
    parser = argparse.ArgumentParser(
        description='3D山岭图生成器 — 从CSV/Excel生成交互式3D山岭图HTML',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python ridge3d.py data.csv
  python ridge3d.py data.csv -t "Meta分析" -a "张三" -c ocean -o output.html
  python ridge3d.py data.xlsx --sheet "Sheet1"

色彩方案: forest, ocean, sunset, glacier, lava, aurora
        """,
    )
    parser.add_argument('input', nargs='?', default=None, help='输入文件路径 (CSV或Excel)')
    parser.add_argument('-t', '--title', default='3D Ridge Plot', help='图表标题')
    parser.add_argument('--subtitle', default='', help='副标题')
    parser.add_argument('-a', '--author', default='', help='署名 (显示在标题下方)')
    parser.add_argument('-c', '--colorscheme', default='forest',
                        choices=list(COLORSCHEMES.keys()), default='golden',
                        help='配色方案 (默认: golden)')
    parser.add_argument('-o', '--output', default='', help='输出HTML路径 (默认: 输入文件名_3D.html)')
    parser.add_argument('--sheet', default=None, help='Excel工作表名')
    parser.add_argument('--xlabel', default='Study / Subgroup', help='X轴标签')
    parser.add_argument('--ylabel', default='Effect Size', help='Y轴标签')
    parser.add_argument('--list-colors', action='store_true', help='列出所有配色方案')

    args = parser.parse_args()

    if args.list_colors:
        print("\n可用的配色方案:\n")
        for key, cs in COLORSCHEMES.items():
            print(f"  {key:12s} — {cs['name']}")
        print()
        return

    if not args.input:
        parser.print_help()
        print("\nError: input file required")
        return

    # Load data
    ext = os.path.splitext(args.input)[1].lower()
    if ext in ('.xlsx', '.xls'):
        rows = load_excel(args.input, args.sheet)
    else:
        rows = load_csv(args.input)

    print(f"Loaded {len(rows)} rows")

    # Parse data
    data = parse_data(rows, args)
    n_groups = len(set(d[5] for d in data if d[5]))
    print(f"Parsed {len(data)} records, {n_groups} groups")

    # Build figure
    fig = build_ridge(data, args)

    # 输出
    if not args.output:
        base = os.path.splitext(os.path.basename(args.input))[0]
        args.output = os.path.join(os.path.dirname(args.input) or '.',
                                    f"{base}_3D山岭图.html")
    fig.write_html(args.output, include_plotlyjs='cdn', full_html=True)
    fsize = os.path.getsize(args.output) / 1024
    print(f"\n[Done] Output: {args.output}")
    print(f"   Data points: {len(data)} | File size: {fsize:.1f} KB")
    print(f"   Colorscheme: {COLORSCHEMES[args.colorscheme]['name']}")
    if args.author:
        print(f"   Author: {args.author}")
    print(f"\n   Open in browser to view (rotate/zoom/hover enabled)")


if __name__ == '__main__':
    main()
